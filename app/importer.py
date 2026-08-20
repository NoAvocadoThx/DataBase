"""Excel 导入/导出、去重与合并。"""
import io, re
from typing import Optional

import openpyxl
from sqlalchemy.orm import Session

from . import history
from .models import DuplicateCandidate, Expert, Participation, Tag, live

EXCEL_COLS = ["姓名", "单位", "职务", "研究方向", "手机", "邮箱", "微信", "简介", "标签", "备注"]
SPLIT = re.compile(r"[,，;；/、\s]+")


def split_tags(text: str) -> list[str]:
    return [t for t in SPLIT.split(text or "") if t]


def get_or_create_tag(s: Session, name: str) -> Tag:
    name = name.strip()
    t = s.query(Tag).filter_by(name=name).first()
    if not t:
        t = Tag(name=name)
        s.add(t)
        s.flush()
    return t


def mask(v: Optional[str]) -> str:
    if not v:
        return ""
    return v[:3] + "****" + v[-2:] if len(v) > 5 else "****"


def register_duplicates(s: Session, e: Expert, reopen: bool = False):
    """与 e 同名但不同 id 的在库专家，登记为疑似重复。reopen=True 时把因合并/删除关闭的对子重新打开（人工标记的“不同人”保留）。"""
    for other in live(s.query(Expert)).filter(Expert.name == e.name, Expert.id != e.id):
        a, b = sorted((e.id, other.id))
        exists = s.query(DuplicateCandidate).filter_by(expert_a_id=a, expert_b_id=b).first()
        if not exists:
            s.add(DuplicateCandidate(expert_a_id=a, expert_b_id=b))
        elif reopen and exists.status in ("merged", "closed"):
            exists.status = "pending"


def upsert_expert(s: Session, data: dict, source: str, actor: str = "", action: str = "import") -> tuple[Expert, bool]:
    """姓名+单位全等 → 更新; 否则新建。返回 (专家, 是否新建)。全部写入修改历史。"""
    name, org = data.get("name", "").strip(), data.get("org", "").strip()
    e = live(s.query(Expert)).filter_by(name=name, org=org).first()
    created = e is None
    if created:
        e = Expert(name=name, org=org)
        s.add(e)
        s.flush()
    before = {} if created else history.snapshot(e)
    for k in ("title", "field", "phone", "email", "wechat", "bio", "note", "source_text"):
        v = (data.get(k) or "").strip()
        if v:
            setattr(e, k, v)
    e.source = source
    if data.get("tags") is not None:
        e.tags = [get_or_create_tag(s, t) for t in data["tags"]]
    if created:
        register_duplicates(s, e)
        history.log(s, actor, action, e, history.snapshot(e), f"新建，来源: {source}")
    else:
        history.log_update(s, actor, e, before, action, f"来源: {source}")
    return e, created


def import_excel(s: Session, data: bytes, filename: str, actor: str = "") -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    if not rows:
        return {"error": "空文件"}
    header = [str(h).strip() if h else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(header)}
    if "姓名" not in idx:
        return {"error": "缺少“姓名”列，模板列名: " + " / ".join(EXCEL_COLS)}
    created = updated = 0

    def cell(r, col):
        i = idx.get(col)
        v = r[i] if i is not None and i < len(r) else None
        return str(v).strip() if v is not None else ""

    for r in rows[1:]:
        if not cell(r, "姓名"):
            continue
        d = dict(name=cell(r, "姓名"), org=cell(r, "单位"), title=cell(r, "职务"),
                 field=cell(r, "研究方向"), phone=cell(r, "手机"), email=cell(r, "邮箱"),
                 wechat=cell(r, "微信"), bio=cell(r, "简介"), note=cell(r, "备注"),
                 tags=split_tags(cell(r, "标签")))
        _, is_new = upsert_expert(s, d, filename, actor, "import")
        created += is_new
        updated += not is_new
    s.commit()
    dup = s.query(DuplicateCandidate).filter_by(status="pending").count()
    return {"created": created, "updated": updated, "pending_dup": dup}


def export_excel(s: Session) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(EXCEL_COLS)
    for e in live(s.query(Expert)).order_by(Expert.id):
        ws.append([e.name, e.org, e.title, e.field, e.phone, e.email, e.wechat, e.bio,
                   ", ".join(t.name for t in e.tags), e.note])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def merge_experts(s: Session, keep: Expert, drop: Expert, actor: str = "") -> Expert:
    """把 drop 合并进 keep: 空字段补齐、标签并集、合作历史迁移；drop 进回收站（可恢复）。"""
    before = history.snapshot(keep)
    for k in ("org", "title", "field", "phone", "email", "wechat", "bio", "note", "source"):
        if not getattr(keep, k) and getattr(drop, k):
            setattr(keep, k, getattr(drop, k))
    names = {t.name for t in keep.tags}
    keep.tags.extend(t for t in drop.tags if t.name not in names)
    for m in list(drop.meetings):
        m.expert = keep
    s.query(DuplicateCandidate).filter(
        (DuplicateCandidate.expert_a_id == drop.id) | (DuplicateCandidate.expert_b_id == drop.id)
    ).update({"status": "merged"}, synchronize_session=False)
    soft_delete(s, drop, actor, f"已合并到 #{keep.id} {keep.name}（{keep.org}）")
    note = f"并入 #{drop.id} {drop.name}（{drop.org}）"
    if not history.log_update(s, actor, keep, before, "merge", note):
        history.log(s, actor, "merge", keep, {}, note)
    s.commit()
    return keep


def soft_delete(s: Session, e: Expert, actor: str, reason: str = ""):
    from datetime import datetime
    e.deleted_at, e.deleted_by = datetime.now(), actor
    s.query(DuplicateCandidate).filter(
        ((DuplicateCandidate.expert_a_id == e.id) | (DuplicateCandidate.expert_b_id == e.id)),
        DuplicateCandidate.status == "pending").update({"status": "closed"}, synchronize_session=False)
    history.log(s, actor, "delete", e, history.snapshot(e), reason)


def restore(s: Session, e: Expert, actor: str):
    e.deleted_at, e.deleted_by = None, ""
    register_duplicates(s, e, reopen=True)
    history.log(s, actor, "restore", e, {}, "从回收站恢复")


def purge(s: Session, e: Expert, actor: str):
    """彻底删除，历史记录保留（不设外键）。"""
    history.log(s, actor, "purge", e, history.snapshot(e), "彻底删除")
    s.query(DuplicateCandidate).filter(
        (DuplicateCandidate.expert_a_id == e.id) | (DuplicateCandidate.expert_b_id == e.id)).delete()
    s.delete(e)
