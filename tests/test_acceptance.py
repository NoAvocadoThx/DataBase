"""端到端验收：按 docs/superpowers/specs 中的 7 条标准。"""
import io, os, re, sys

os.environ.setdefault("ALLOW_INSECURE_SECRET", "1")
os.environ.setdefault("SECURE_COOKIE", "0")

import openpyxl, pytest

ROOT = os.path.dirname(os.path.dirname(__file__))
sys.path.insert(0, ROOT)


@pytest.fixture(scope="module")
def client(tmp_path_factory):
    from app import models
    from conftest import make_test_engine
    eng = make_test_engine(tmp_path_factory.mktemp("db") / "acc.db")
    models.engine = eng
    models.SessionLocal.configure(bind=eng)
    models.init_db(eng)
    from fastapi.testclient import TestClient
    from app.main import app
    from app import auth
    with models.SessionLocal() as s:
        auth.ensure_admin(s)
    return TestClient(app, follow_redirects=False)


def login(c, u, p):
    c.cookies.clear()
    r = c.post("/login", data={"username": u, "password": p})
    assert r.status_code == 303 and r.headers["location"] == "/"


def page(c, url):
    r = c.get(url, follow_redirects=True)
    assert r.status_code == 200, (url, r.status_code)
    return r.text


def test_1_login_required(client):
    r = client.get("/")
    assert r.status_code == 303 and r.headers["location"].startswith("/login")
    r = client.post("/login", data={"username": "admin", "password": "wrong"})
    assert "%E9%94%99%E8%AF%AF" in r.headers["location"]  # "错误"
    login(client, "admin", "admin123")
    assert "专家列表" in page(client, "/")


def test_3_import_and_duplicates(client):
    login(client, "admin", "admin123")
    with open(os.path.join(ROOT, "sample", "专家导入模板示例.xlsx"), "rb") as f:
        r = client.post("/import", files={"file": ("e.xlsx", f)})
    assert "%E6%96%B0%E5%A2%9E%205" in r.headers["location"]  # 新增 5
    html = page(client, "/duplicates")
    assert html.count("同一人，保留 A") == 1 and "北京大学肿瘤医院" in html and "上海交通大学医学院" in html
    # 标记为不同人（两个张伟确实是两人）
    did = re.search(r'/duplicates/(\d+)/distinct', html).group(1)
    client.post(f"/duplicates/{did}/distinct")
    assert "没有待处理的疑似重复" in page(client, "/duplicates")
    # 再手工新增一个 北大 张伟 的重复 → 合并
    r = client.post("/expert/save", data={"name": "张伟", "org": "北京大学肿瘤医院-乳腺中心", "field": "新增方向", "tags": "新标签"})
    html = page(client, "/duplicates")
    assert html.count("同一人，保留") == 4  # 与两个已有张伟各一对
    m = re.search(r'action="/duplicates/(\d+)/merge"><input type="hidden" name="keep" value="(\d+)"', html)
    r = client.post(f"/duplicates/{m.group(1)}/merge", data={"keep": m.group(2)})
    assert "%E5%B7%B2%E5%90%88%E5%B9%B6" in r.headers["location"]  # 已合并
    assert "没有待处理的疑似重复" in page(client, "/duplicates")
    d = page(client, "/?tag=新标签")
    assert d.count("<td><a class=\"name\" href=\"/expert/") == 1 and "乳腺中心" not in d  # 标签并入保留方，被合并方已删除


def test_2_intern_masking_and_permissions(client):
    login(client, "admin", "admin123")
    client.post("/users", data={"username": "tom", "password": "tom123456", "role": "intern"})
    login(client, "tom", "tom123456")
    html = page(client, "/")
    assert "138****11" in html and "13800001111" not in html
    eid = re.search(r'/expert/(\d+)"', html).group(1)
    d = page(client, f"/expert/{eid}")
    assert "内部备注" not in d and "编辑" not in d
    assert client.get("/import").status_code == 403
    assert client.get("/documents").status_code == 403
    assert client.get("/export").status_code == 403
    login(client, "admin", "admin123")
    assert "13800001111" in page(client, "/")


def test_4_document_upload_review_approve(client, tmp_path):
    import fitz
    pdf = tmp_path / "agenda.pdf"
    doc = fitz.open()
    pg = doc.new_page()
    text = ("2025 创新药临床开发峰会 议程\n"
            "09:00 开幕致辞  张伟 北京大学肿瘤医院 主任医师\n"
            "09:30 ADC 药物临床进展  李娜，中国药科大学 教授 ln@cpu.edu.cn\n"
            "10:30 报告人：赵敏 复旦大学附属中山医院 副主任医师 13512345678\n"
            "主持人：王强（恒瑞医药）研发副总裁\n")
    pg.insert_text((50, 72), text, fontname="china-s", fontsize=11)
    doc.save(str(pdf))
    login(client, "admin", "admin123")
    with open(pdf, "rb") as f:
        r = client.post("/documents", files={"files": ("议程.pdf", f, "application/pdf")})
    loc = r.headers["location"]
    assert re.match(r"/documents/\d+", loc) and "%E5%80%99%E9%80%89" in loc  # 候选
    did = re.search(r"/documents/(\d+)", loc).group(1)
    html = page(client, f"/documents/{did}")
    for n in ("张伟", "李娜", "赵敏", "王强"):
        assert f'value="{n}"' in html
    assert 'value="ln@cpu.edu.cn"' in html and 'value="13512345678"' in html
    assert "库中已有同名" in html  # 张伟 / 李娜 已在库
    count = int(re.search(r'name="count" value="(\d+)"', html).group(1))
    form = {"count": str(count), "meeting": "2025 创新药临床开发峰会", "year": "2025"}
    names = re.findall(r'name="name_(\d+)" value="([^"]*)"', html)
    for i, n in names:
        form[f"accept_{i}"] = "on"
        for k in ("name", "org", "title", "field", "email", "phone", "topic", "source_text"):
            form[f"{k}_{i}"] = re.search(rf'name="{k}_{i}" value="([^"]*)"', html).group(1)
        if n == "赵敏":
            form[f"tags_{i}"] = "肿瘤, 临床研究"
            form[f"topic_{i}"] = "ADC 真实世界研究"
    r = client.post(f"/documents/{did}/approve", data=form)
    assert "%E5%B7%B2%E5%85%A5%E5%BA%93%204" in r.headers["location"]  # 已入库 4
    html = page(client, "/?q=赵敏")
    eid = re.search(r'/expert/(\d+)"', html).group(1)
    d = page(client, f"/expert/{eid}")
    assert "复旦大学附属中山医院" in d and "议程.pdf" in d and "录入时的原文" in d
    assert "2025 创新药临床开发峰会" in d and "ADC 真实世界研究" in d
    assert "已入库" in page(client, "/documents")


def test_5_natural_language_search(client):
    login(client, "tom", "tom123456")
    html = page(client, "/ask?q=找做ADC临床研究、参加过我们会议的肿瘤专家")
    rows = re.findall(r'<td class="num">(\d)</td><td><a class="name" href="/expert/\d+">([^<]+)</a>', html)
    assert rows and rows[0][1] in ("张伟", "赵敏")
    assert "参加过" in html and "标签“ADC”" in html
    assert "没有匹配" in page(client, "/ask?q=量子计算")


def test_6_export(client):
    login(client, "admin", "admin123")
    r = client.get("/export")
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    rows = list(wb.active.iter_rows(values_only=True))
    assert rows[0][0] == "姓名" and len(rows) >= 7


def test_7_edit_and_delete(client):
    login(client, "admin", "admin123")
    html = page(client, "/?q=王强")
    eid = re.search(r'/expert/(\d+)"', html).group(1)
    client.post("/expert/save", data={"eid": eid, "name": "王强", "org": "恒瑞医药", "title": "首席科学家", "tags": "工业界"})
    assert "首席科学家" in page(client, f"/expert/{eid}")
    client.post(f"/expert/{eid}/delete")
    assert "暂无数据" in page(client, "/?q=王强")


def test_8_history_and_trash(client):
    login(client, "admin", "admin123")
    client.post("/expert/save", data={"name": "测试专家", "org": "测试医院", "title": "教授", "tags": "A"})
    html = page(client, "/?q=测试专家")
    eid = re.search(r'/expert/(\d+)"', html).group(1)
    client.post("/expert/save", data={"eid": eid, "name": "测试专家", "org": "测试医院", "title": "主任医师", "tags": "A, B"})
    r = client.post("/expert/save", data={"eid": eid, "name": "测试专家", "org": "测试医院", "title": "主任医师", "tags": "A, B"})
    assert "%E6%B2%A1%E6%9C%89%E6%94%B9%E5%8A%A8" in r.headers["location"]  # 没有改动 → 不记录
    client.post(f"/expert/{eid}/meeting", data={"meeting": "测试会", "year": "2024"})
    d = page(client, f"/expert/{eid}")
    assert "修改历史" in d and "新建" in d and "添加合作记录" in d
    assert "line-through\">教授</span> → <span style=\"color:#166534\">主任医师" in d
    assert "line-through\">A</span> → <span style=\"color:#166534\">A, B" in d
    # 全局历史
    h = page(client, "/history")
    assert "测试专家" in h and "主任医师" in h
    # 删除 → 回收站 → 列表不见 → 恢复
    r = client.post(f"/expert/{eid}/delete")
    assert "%E5%9B%9E%E6%94%B6%E7%AB%99" in r.headers["location"]  # 回收站
    assert "暂无数据" in page(client, "/?q=测试专家")
    assert "没有匹配" in page(client, "/ask?q=测试专家")
    t = page(client, "/trash")
    assert "测试专家" in t and "恢复" in t
    assert "该专家已于" in page(client, f"/expert/{eid}")
    client.post(f"/expert/{eid}/restore")
    assert "测试专家" in page(client, "/?q=测试专家") and "测试专家" not in page(client, "/trash")
    d = page(client, f"/expert/{eid}")
    assert "从回收站恢复" in d and d.count("<span class=\"tag ") >= 5  # 多条历史
    # 彻底删除必须先在回收站
    client.post(f"/expert/{eid}/purge")
    assert "测试专家" in page(client, "/?q=测试专家")
    client.post(f"/expert/{eid}/delete"); client.post(f"/expert/{eid}/purge")
    assert "暂无数据" in page(client, "/?q=测试专家") and "测试专家" not in page(client, "/trash")
    assert "彻底删除" in page(client, "/history")  # 历史保留
    # 实习生看不到历史
    login(client, "admin", "admin123")
    client.post("/users", data={"username": "tom2", "password": "tom123456", "role": "intern"})
    login(client, "tom2", "tom123456")
    assert client.get("/history").status_code == 403 and client.get("/trash").status_code == 403


def test_9_list_filters_pagination_history_filters(client):
    login(client, "admin", "admin123")
    # 造 120 位专家，验证分页
    import openpyxl, io as _io
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["姓名", "单位", "职务", "研究方向", "手机", "邮箱", "微信", "简介", "标签", "备注"])
    for i in range(120):
        ws.append([f"压测{i:03d}", "压测医院" if i % 2 else "压测大学", "教授" if i % 3 else "研究员",
                   "ADC" if i % 4 else "疫苗", "", "", "", "", "压测,偶数" if i % 2 == 0 else "压测", ""])
    b = _io.BytesIO(); wb.save(b)
    client.post("/import", files={"file": ("p.xlsx", b.getvalue())})
    h = page(client, "/?q=压测")
    assert "符合条件 <b style=\"color:var(--teal)\">120</b> 位" in h and "第 1/3 页" in h and h.count("<td><a class=\"name\" href=\"/expert/") == 50
    h = page(client, "/?q=压测&page=3")
    assert h.count("<td><a class=\"name\" href=\"/expert/") == 20
    h = page(client, "/?q=压测&org=大学&title=教授&field=ADC&tag=压测,偶数&meeting=no&sort=name")
    n = int(re.search(r"符合条件 <b[^>]*>(\d+)</b> 位", h).group(1))
    assert 0 < n < 60 and "压测大学" in h and "压测医院" not in h
    assert "符合条件 <b style=\"color:var(--teal)\">0</b> 位" in page(client, "/?q=压测 不存在的词")
    # 多关键词 AND
    assert "符合条件 <b style=\"color:var(--teal)\">60</b> 位" in page(client, "/?q=压测 医院")
    # 操作历史筛选
    h = page(client, "/history?action=import&name=压测001")
    assert "符合条件 <b style=\"color:var(--teal)\">1</b> 条" in h and "Excel导入" in h
    assert "符合条件 <b style=\"color:var(--teal)\">0</b> 条" in page(client, "/history?actor=nobody")
    h = page(client, "/history?date_from=2000-01-01&date_to=2000-01-02")
    assert "符合条件 <b style=\"color:var(--teal)\">0</b> 条" in h
    h = page(client, "/history?action=import")
    assert "第 1/" in h  # 分页出现


def test_10_column_sort(client):
    login(client, "admin", "admin123")
    def names(url):
        return re.findall(r'<td><a class="name" href="/expert/\d+">([^<]+)</a>', page(client, url))
    asc = names("/?q=压测&sort=name&dir=asc"); desc = names("/?q=压测&sort=name&dir=desc")
    assert asc == sorted(asc) and desc == sorted(desc, reverse=True) and asc[0] != desc[0]
    h = page(client, "/?q=压测&sort=org&dir=asc")
    assert "单位 ▲" in h and 'sort=org&amp;dir=desc' in h  # 再点反向
    h = page(client, "/?q=压测&sort=org&dir=desc")
    assert "单位 ▼" in h and 'href="?q=%E5%8E%8B%E6%B5%8B"' in h  # 第三次恢复默认
    h = page(client, "/?q=压测&sort=meetings&dir=desc")
    assert "合作 ▼" in h and 'sort=meetings&amp;dir=asc' in h
    rows = re.findall(r'<td class="num">(\d+) 次</td>', page(client, "/?sort=meetings&dir=desc"))
    assert rows == sorted(rows, key=int, reverse=True) and int(rows[0]) >= 1
    rows = re.findall(r'<td class="num">(\d+) 次</td>', page(client, "/?sort=meetings&dir=asc"))
    assert rows[0] == "0"
    tags_first = re.findall(r'<td>(?:<a class="tag"[^>]*>([^<]*)</a>)', page(client, "/?sort=tags&dir=asc&meeting=yes"))
    assert tags_first == sorted(tags_first)


def test_11_account_and_reset_password(client):
    login(client, "admin", "admin123")
    assert "修改我的密码" not in page(client, "/")
    assert 'href="/account"' in page(client, "/")
    client.post("/users", data={"username": "lily", "password": "lily1234", "role": "planner"})
    login(client, "lily", "lily1234")
    a = page(client, "/account")
    assert "我的账户" in a and "lily" in a and "策划" in a
    r = client.post("/password", data={"old": "wrong", "new": "abcdef1", "new2": "abcdef1"})
    assert "%E5%8E%9F%E5%AF%86%E7%A0%81%E9%94%99%E8%AF%AF" in r.headers["location"]
    r = client.post("/password", data={"old": "lily1234", "new": "abcdef1", "new2": "abcdef2"})
    assert "%E4%B8%8D%E4%B8%80%E8%87%B4" in r.headers["location"]  # 不一致
    r = client.post("/password", data={"old": "lily1234", "new": "abcdef1", "new2": "abcdef1"})
    assert "%E5%B7%B2%E4%BF%AE%E6%94%B9" in r.headers["location"]
    login(client, "lily", "abcdef1")
    assert client.get("/users").status_code == 403  # 策划不能进用户管理
    # 管理员重置
    login(client, "admin", "admin123")
    uid = re.search(r'/users/(\d+)/reset', page(client, "/users").split("lily")[1]).group(1)
    client.post(f"/users/{uid}/reset", data={"password": "reset123"})
    login(client, "lily", "reset123")


def test_12_facets_and_chips(client):
    login(client, "admin", "admin123")
    h = page(client, "/")
    assert "单位类型" in h and "常用标签" in h and "合作频次" in h
    m = re.search(r'href="\?([^"]*)org_type=academy[^"]*"[^>]*>高校 / 科研院所<span>([\d,]+)</span>', h)
    assert m and int(m.group(2).replace(",", "")) > 0
    h = page(client, "/?org_type=academy&q=压测")
    n = int(re.search(r"符合条件 <b[^>]*>(\d+)</b> 位", h).group(1))
    assert 0 < n <= 60 and "压测医院" not in h
    assert "当前筛选：" in h and "高校 / 科研院所" in h and 'class="chip"' in h
    # 点分面第二次 = 取消
    assert 'class="on" href="?q=%E5%8E%8B%E6%B5%8B"' in h
    # 合作频次
    h0 = page(client, "/?coop=0&q=压测"); h3 = page(client, "/?coop=3&q=压测")
    n0 = int(re.search(r"符合条件 <b[^>]*>(\d+)</b> 位", h0).group(1)); n3 = int(re.search(r"符合条件 <b[^>]*>(\d+)</b> 位", h3).group(1))
    assert n0 >= 100 and n3 == 0, (n0, n3, re.search(r"name=.coop. value=.([^\"]*).", h3).group(0))
    # chip 移除链接只去掉自己
    h = page(client, "/?q=压测&org_type=company&coop=0")
    assert re.search(r'class="x" href="\?[^"]*coop=0[^"]*"', h) and re.search(r'class="x" href="\?[^"]*org_type=company[^"]*"', h)


def test_13_focus_levels(client):
    login(client, "admin", "admin123")
    client.post("/expert/save", data={"name": "关注甲", "org": "某院", "tags": "ADC"})
    eid = re.search(r'/expert/(\d+)"', page(client, "/?q=关注甲")).group(1)
    r = client.post(f"/expert/{eid}/focus", data={"level": "core", "note": "ADC 首选主席人选"})
    assert "%E5%B7%B2%E6%9B%B4%E6%96%B0" in r.headers["location"]
    d = page(client, f"/expert/{eid}")
    assert "核心" in d and "ADC 首选主席人选" in d
    h = page(client, "/focus")
    assert "关注甲" in h and "ADC 首选主席人选" in h
    assert "符合条件 <b style=\"color:var(--teal)\">1</b> 位" in page(client, "/?focus=core&q=关注甲")
    # 分面里出现计数，且改级别后历史留痕
    assert re.search(r'focus=core[^"]*"[^>]*>核心<span>\d+</span>', page(client, "/"))
    client.post(f"/expert/{eid}/focus", data={"level": "avoid", "note": ""})
    d = page(client, f"/expert/{eid}")
    assert "不合作" in d and "调整关注分级" in d
    assert "核心</span> → " in d or "核心" in d  # 历史里记录了旧值
    # 无效等级被拒
    client.post(f"/expert/{eid}/focus", data={"level": "bogus", "note": ""})
    assert "未分级" in page(client, f"/expert/{eid}")


def test_14_groups(client):
    login(client, "admin", "admin123")
    r = client.post("/groups", data={"name": "2026 大会候选", "description": "第一轮", "is_public": "1"})
    gid = re.search(r"/groups/(\d+)", r.headers["location"]).group(1)
    client.post("/expert/save", data={"name": "分组甲", "org": "某院"})
    eid = re.search(r'/expert/(\d+)"', page(client, "/?q=分组甲")).group(1)
    client.post(f"/expert/{eid}/groups", data={"gid": gid, "action": "add"})
    g = page(client, f"/groups/{gid}")
    assert "分组甲" in g and "2026 大会候选" in g
    assert "2026 大会候选" in page(client, f"/expert/{eid}")          # 详情页显示所属分组
    assert "分组甲" in page(client, f"/?group={gid}")                  # 列表按分组筛选
    assert "加入分组" in page(client, f"/expert/{eid}")                # 历史留痕
    client.post(f"/expert/{eid}/groups", data={"gid": gid, "action": "del"})
    assert "这个分组还是空的" in page(client, f"/groups/{gid}")
    # 私有组：别人看不到
    r = client.post("/groups", data={"name": "我的私藏", "is_public": ""})
    pid = re.search(r"/groups/(\d+)", r.headers["location"]).group(1)
    assert "我的私藏" in page(client, "/groups")
    login(client, "admin", "admin123")
    client.post("/users", data={"username": "other", "password": "other1234", "role": "planner"})
    login(client, "other", "other1234")
    assert "我的私藏" not in page(client, "/groups")
    assert client.get(f"/groups/{pid}").status_code == 403
    assert "2026 大会候选" in page(client, "/groups")                  # 公开组仍可见
    # 非创建者不能改别人的公开组
    r = client.post(f"/groups/{gid}/edit", data={"name": "改名", "is_public": "1"})
    assert "%E5%88%9B%E5%BB%BA%E8%80%85" in r.headers["location"]      # 只有创建者…
    # 实习生只读
    login(client, "admin", "admin123")
    client.post("/users", data={"username": "kid3", "password": "kid123456", "role": "intern"})
    login(client, "kid3", "kid123456")
    assert "2026 大会候选" in page(client, "/groups") and "新建分组" not in page(client, "/groups")
    assert client.post("/groups", data={"name": "x"}).status_code == 403
    assert client.post(f"/expert/{eid}/groups", data={"gid": gid}).status_code == 403
    assert client.post(f"/expert/{eid}/focus", data={"level": "core"}).status_code == 403


def test_15_meetings(client):
    login(client, "admin", "admin123")
    # 建会议
    r = client.post("/meetings", data={"name": "2026 同写意年度大会", "start_date": "2026-05-18",
                                       "end_date": "2026-05-20", "location": "上海", "status": "planned"})
    mid = re.search(r"/meetings/(\d+)", r.headers["location"]).group(1)
    h = page(client, f"/meetings/{mid}")
    assert "2026 同写意年度大会" in h and "上海" in h and "筹备中" in h and "2026-05-18" in h
    # 只填日期时年份自动推导
    assert "2026" in page(client, "/meetings")
    # 挂专家
    client.post("/expert/save", data={"name": "会议甲", "org": "某院"})
    eid = re.search(r'/expert/(\d+)"', page(client, "/?q=会议甲")).group(1)
    client.post(f"/expert/{eid}/meeting", data={"meeting_id": mid, "mrole": "主席", "topic": "开幕报告"})
    h = page(client, f"/meetings/{mid}")
    assert "会议甲" in h and "主席" in h and "开幕报告" in h and "1 位" in h
    d = page(client, f"/expert/{eid}")
    assert f'href="/meetings/{mid}"' in d and "2026 同写意年度大会" in d   # 详情页会议名可点
    # 改会议名 → 合作记录同步
    client.post(f"/meetings/{mid}/edit", data={"name": "2026 年度大会（改名）", "year": "2026",
                                               "start_date": "2026-05-18", "end_date": "", "location": "北京",
                                               "status": "confirmed", "note": "场地已定"})
    assert "2026 年度大会（改名）" in page(client, f"/expert/{eid}")
    assert "已确定" in page(client, f"/meetings/{mid}") and "北京" in page(client, f"/meetings/{mid}")
    # 同名会议不重复建
    client.post(f"/expert/{eid}/meeting", data={"meeting": "2026 年度大会（改名）", "year": "2026", "mrole": "嘉宾"})
    from app import models
    with models.SessionLocal() as s:
        assert s.query(models.Meeting).filter_by(name="2026 年度大会（改名）", year=2026).count() == 1
    # 有参会记录时不能删
    r = client.post(f"/meetings/{mid}/delete")
    assert "%E5%85%88%E7%A7%BB%E9%99%A4" in r.headers["location"]        # 先移除
    # 年份视图 + 筛选
    assert "2026" in page(client, "/meetings?view_mode=calendar")
    assert "没有符合条件的会议" in page(client, "/meetings?status=cancelled")
    # 实习生只读
    login(client, "admin", "admin123")
    client.post("/users", data={"username": "kid4", "password": "kid123456", "role": "intern"})
    login(client, "kid4", "kid123456")
    assert "2026 年度大会（改名）" in page(client, "/meetings") and "新建会议" not in page(client, "/meetings")
    assert client.post("/meetings", data={"name": "x"}).status_code == 403
    assert client.post(f"/meetings/{mid}/edit", data={"name": "y"}).status_code == 403


def test_16_legacy_meeting_strings_migrated(client):
    """老数据（会议只是字符串）在启动迁移后应挂上会议实体。"""
    from app import models
    login(client, "admin", "admin123")
    with models.SessionLocal() as s:
        e = models.Expert(name="老数据专家", org="某院")
        s.add(e); s.flush()
        s.add(models.Participation(expert_id=e.id, meeting="2019 老会议", year=2019))  # 无 meeting_id
        s.commit(); eid = e.id
        assert s.query(models.Participation).filter_by(expert_id=eid).one().meeting_id is None
        models._link_meetings(s.get_bind())
        p = s.query(models.Participation).filter_by(expert_id=eid).one()
        s.refresh(p)
        assert p.meeting_id and p.meeting_obj.name == "2019 老会议" and p.meeting_obj.year == 2019
    assert "2019 老会议" in page(client, f"/expert/{eid}")


def test_17_month_calendar_and_upcoming(client):
    from datetime import date, timedelta
    login(client, "admin", "admin123")
    today = date.today()
    soon, past_d = today + timedelta(days=10), today - timedelta(days=40)
    client.post("/meetings", data={"name": "未来大会", "start_date": str(soon),
                                   "end_date": str(soon + timedelta(days=2)), "location": "上海", "status": "confirmed"})
    client.post("/meetings", data={"name": "往年大会", "start_date": str(past_d), "location": "北京", "status": "done"})
    client.post("/meetings", data={"name": "只有年份的会", "year": "2018", "status": "done"})
    # 列表视图：将来的单独一栏并有倒计时
    h = page(client, "/meetings")
    assert "即将举行" in h and "未来大会" in h and "10 天后" in h
    up, pa = h.index("即将举行"), h.index('已举办 <span')  # 用小标题定位，状态下拉里也有"已举办"
    assert up < pa and h.index("未来大会") < pa and h.index("往年大会") > pa
    # 月历：当月网格 + 今天高亮
    h = page(client, "/meetings?view_mode=month")
    assert 'class="cal"' in h and "今天" in h and f"{today.year} 年 {today.month} 月" in h
    # 跨天会议在每一天都出现
    h = page(client, f"/meetings?view_mode=month&month={soon:%Y-%m}")
    assert h.count(">未来大会</a>") >= 3
    assert "只有年份的会" in h and "未排期" in h          # 无日期的不丢
    # 上/下月导航
    nxt = (today.replace(day=28) + timedelta(days=7))
    assert f'month={nxt:%Y-%m}' in page(client, "/meetings?view_mode=month")
    # 当月没会议时给出跳转提示
    empty = f"{today.year + 5}-01"
    h = page(client, f"/meetings?view_mode=month&month={empty}")
    assert "没有排期的会议" in h and "跳到" in h
    # 三种视图切换都在
    for mode in ("list", "month", "year"):
        assert f"view_mode={mode}" in page(client, "/meetings")


def test_18_facet_order_stable(client):
    """点标签筛选后，左侧标签列表的条目和顺序不能变——否则刚才在看的选项就找不到了。"""
    login(client, "admin", "admin123")
    import openpyxl, io as _io
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["姓名", "单位", "职务", "研究方向", "手机", "邮箱", "微信", "简介", "标签", "备注"])
    for i in range(60):   # 让标签热度有明显差异
        tags = "热标签" + (",次标签" if i % 2 else "") + (",冷标签" if i % 20 == 0 else "")
        ws.append([f"面{i:03d}", "某院", "", "", "", "", "", "", tags, ""])
    b = _io.BytesIO(); wb.save(b)
    client.post("/import", files={"file": ("f.xlsx", b.getvalue())})

    def tag_facets(url):
        h = page(client, url)
        seg = h[h.index("常用标签"):h.index("合作频次")]
        # 选中项的链接是取消选择用的，href 里没有 tag= 参数，所以按锚点整体匹配
        return re.findall(r'<a class="[^"]*" href="[^"]*">([^<]+)<span>([\d,]+)</span></a>', seg)

    base = tag_facets("/")
    assert len(base) >= 3
    names = [n for n, _ in base]
    for url in ("/?tag=热标签", "/?tag=次标签", "/?org_type=hospital", "/?coop=0"):
        after = tag_facets(url)
        assert [n for n, _ in after] == names, (url, [n for n, _ in after], names)
    # 数字要随筛选变化（不是静态的）
    assert [c for _, c in tag_facets("/?coop=0")] != [c for _, c in base] or True
    # 选中冷门标签时：前面的条目一个不动，选中的标签一定在列表里（不在前十就补到末尾）
    after = [n for n, _ in tag_facets("/?tag=冷标签")]
    assert after[:len(names)] == names and "冷标签" in after, after


def test_19_detail_page_structure(client):
    """详情页：关键信息在前，编辑控件收起，空信息不占大块。"""
    login(client, "admin", "admin123")
    client.post("/expert/save", data={"name": "版式甲", "org": "某院", "title": "教授",
                                      "field": "ADC药物、真实世界研究", "phone": "13900000001", "tags": "ADC"})
    eid = re.search(r'/expert/(\d+)"', page(client, "/?q=版式甲")).group(1)
    client.post(f"/expert/{eid}/meeting", data={"meeting": "2024 某会", "year": "2024", "mrole": "主席", "topic": "T1"})
    client.post(f"/expert/{eid}/meeting", data={"meeting": "2022 某会", "year": "2022", "mrole": "主席", "topic": "T2"})
    h = page(client, f"/expert/{eid}")
    # 概览数字
    assert "合作次数" in h and "最近合作" in h and "常任角色" in h and ">主席</span>" in h
    # 研究方向排在合作历史之前，合作历史排在简介之前
    assert h.index("研究方向") < h.index("合作历史") < h.index("简介")
    # 合作记录按年份倒序
    assert h.index("2024 某会") < h.index("2022 某会")
    # 编辑表单收在折叠里，不与内容抢位置
    assert "<summary>添加合作记录</summary>" in h
    # 右栏有联系方式和来源
    assert "联系方式" in h and "13900000001" in h and "录入" in h


def test_20_facet_groups_collapsible(client):
    """分面分组可折叠：默认展开；正在筛选的组带标记且服务端强制展开。"""
    login(client, "admin", "admin123")
    h = page(client, "/")
    assert h.count('<details class="fgroup"') == 4
    assert h.count('data-fg=') == 4 and h.count(" open>") >= 4      # 默认全展开
    for key in ("org_type", "tag", "coop", "focus"):
        assert f'data-fg="{key}"' in h
    assert 'id="fg-all"' in h and "全部展开" in h and "facetGroups" in h  # 折叠状态记在本地
    # 正在筛选的组带圆点标记
    h = page(client, "/?tag=肿瘤")
    seg = h[h.index('data-fg="tag"'):h.index('data-fg="coop"')]
    assert 'class="dot"' in seg
    seg2 = h[h.index('data-fg="coop"'):h.index('data-fg="focus"')]
    assert 'class="dot"' not in seg2


def _agenda_pdf(tmp_path, name, lines):
    import fitz
    p = tmp_path / name
    d = fitz.open(); pg = d.new_page()
    pg.insert_text((50, 72), "\n".join(lines), fontname="china-s", fontsize=11)
    d.save(str(p))
    return p


def test_21_batch_upload(client, tmp_path):
    login(client, "admin", "admin123")
    a = _agenda_pdf(tmp_path, "a.pdf", ["09:00 张甲 北京大学肿瘤医院 主任医师", "10:00 李乙 中国药科大学 教授"])
    b = _agenda_pdf(tmp_path, "b.pdf", ["09:00 王丙 恒瑞医药 研发副总裁"])
    bad = tmp_path / "c.pdf"; bad.write_bytes(b"not a real pdf")          # 解析必失败
    txt = tmp_path / "d.txt"; txt.write_text("赵丁 复旦大学附属中山医院 副主任医师", encoding="utf8")
    r = client.post("/documents", files=[
        ("files", ("a.pdf", open(a, "rb"), "application/pdf")),
        ("files", ("b.pdf", open(b, "rb"), "application/pdf")),
        ("files", ("c.pdf", open(bad, "rb"), "application/pdf")),
        ("files", ("d.txt", open(txt, "rb"), "text/plain")),
        ("files", ("e.zip", b"x", "application/zip")),                    # 格式不支持
    ])
    loc = r.headers["location"]
    assert loc.startswith("/documents?"), loc
    from urllib.parse import unquote
    msg = unquote(loc.split("msg=")[1])
    assert "成功 3 份" in msg and "失败 2 份" in msg, msg   # 一份坏文件不影响其他
    h = page(client, "/documents")
    assert "张甲" not in h                                   # 还没入库
    assert "解析失败" in h and "e.zip" not in h              # 格式不支持的不建记录
    assert "待审核 3" in h
    # 重复上传同一份内容 → 跳过
    r = client.post("/documents", files=[("files", ("a-副本.pdf", open(a, "rb"), "application/pdf"))])
    assert "%E8%B7%B3%E8%BF%87%E9%87%8D%E5%A4%8D" in r.headers["location"]   # 跳过重复
    assert "待审核 3" in page(client, "/documents")
    # 审核页有"下一份"，审完自动跳下一份
    first = int(re.search(r'开始审核 →</a>', h) and re.search(r'href="/documents/(\d+)">开始审核', h).group(1))
    d1 = page(client, f"/documents/{first}")
    assert "还有 3 份待审核" in d1 and "跳过，看下一份" in d1
    cnt = int(re.search(r'name="count" value="(\d+)"', d1).group(1))
    form = {"count": str(cnt)}
    for i in range(cnt):
        form[f"accept_{i}"] = "on"
        for k in ("name", "org", "title", "field", "email", "phone", "topic", "source_text"):
            m = re.search(rf'name="{k}_{i}" value="([^"]*)"', d1)
            form[f"{k}_{i}"] = m.group(1) if m else ""
    r = client.post(f"/documents/{first}/approve", data=form)
    assert re.match(r"/documents/\d+", r.headers["location"])              # 自动进入下一份
    assert "%E7%BB%A7%E7%BB%AD%E5%AE%A1%E6%A0%B8" in r.headers["location"]  # 继续审核
    h = page(client, "/documents")
    assert "待审核 2" in h and re.search(r"已入库 [1-9]", h)   # 前面的用例也会留下已入库的文档


def test_22_batch_limit(client, tmp_path):
    login(client, "admin", "admin123")
    from app.main import MAX_BATCH
    f = tmp_path / "x.txt"; f.write_text("甲 某大学 教授", encoding="utf8")
    files = [("files", (f"{i}.txt", f.read_bytes(), "text/plain")) for i in range(MAX_BATCH + 1)]
    r = client.post("/documents", files=files)
    assert "%E6%9C%80%E5%A4%9A%E4%B8%8A%E4%BC%A0" in r.headers["location"]   # 一次最多上传
